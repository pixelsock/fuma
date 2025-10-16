import argparse, os, io, re, sys
from urllib.parse import urljoin
import requests
from bs4 import BeautifulSoup
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

UA = 'pastel-exporter/1.1 (+github.com/pixelsock/fuma)'

def fetch(url):
    r = requests.get(url, headers={'User-Agent': UA}, timeout=30)
    r.raise_for_status()
    return r

def textify(el):
    return re.sub(r'\s+', ' ', el.get_text(' ', strip=True)).strip()

def normalize_png(data):
    im = Image.open(io.BytesIO(data)).convert('RGBA')
    out = io.BytesIO()
    im.save(out, format='PNG')
    return out.getvalue(), im.width, im.height

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def fit_image(ws, xlimg, row, col, max_w_px=480, max_h_px=320, min_row_h_px=110):
    w, h = xlimg.width, xlimg.height
    scale = min(max_w_px / w, max_h_px / h, 1.0)
    xlimg.width = int(w * scale)
    xlimg.height = int(h * scale)
    ws.add_image(xlimg, f'{get_column_letter(col)}{row}')
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, min_row_h_px * 0.75)

def parse_pastel_share(html, base_url):
    soup = BeautifulSoup(html, 'lxml')
    comments = []
    candidates = soup.select('[data-testid*="comment"], [class*="comment"], [class*="Comment"], article, li')
    seen = set()

    for i, node in enumerate(candidates, 1):
        text = textify(node)
        if not text:
            continue

        # screenshot inside this node
        img = node.select_one('img')
        img_url = None
        if img and img.get('src'):
            img_url = img['src']
            if img_url.startswith('/'):
                img_url = urljoin(base_url, img_url)

        # author / date (best-effort)
        author, created_at = None, None
        m = re.search(r'([A-Z][a-z]+(?: [A-Z][a-z]+)+)\s*[·|•-]\s*([A-Za-z]{3,}\s+\d{1,2}(?:,\s*\d{4})?)', text)
        if m:
            author, created_at = m.group(1), m.group(2)
        else:
            m2 = re.search(r'([A-Za-z]{3,}\s+\d{1,2}(?:,\s*\d{4})?)', text)
            if m2:
                created_at = m2.group(1)

        cleaned = re.sub(r'\b(Reply|Resolve|Unresolve|Copy link|Share|Edit|Delete)\b', '', text, flags=re.I).strip()
        cleaned = re.sub(r'\s{2,}', ' ', cleaned)

        sig = (cleaned[:140], img_url or '')
        if sig in seen:
            continue
        seen.add(sig)

        comments.append({
            'id': f'c{i}',
            'author': author or 'Reviewer',
            'date': created_at or '',
            'page_title': '',
            'page_url': base_url,
            'text': cleaned,
            'screenshot_url': img_url
        })
    return comments

def main():
    ap = argparse.ArgumentParser(description='Export Pastel comments (public board) to Excel with embedded screenshots.')
    ap.add_argument('url', help='Public Pastel share URL')
    ap.add_argument('--out', default='pastel_comments.xlsx', help='Output Excel filename')
    args = ap.parse_args()

    os.makedirs('pastel_screenshots', exist_ok=True)

    html = fetch(args.url).text
    comments = parse_pastel_share(html, args.url)
    if not comments:
        print('No comments found. If Pastel UI changed, selectors may need tweaks.')
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pastel Comments'

    headers = ['ID', 'Author', 'Date', 'Page Title', 'Page URL', 'Comment', 'Screenshot']
    ws.append(headers)

    # column widths
    set_col_width(ws, 1, 10)
    set_col_width(ws, 2, 20)
    set_col_width(ws, 3, 18)
    set_col_width(ws, 4, 42)
    set_col_width(ws, 5, 45)
    set_col_width(ws, 6, 90)
    set_col_width(ws, 7, 62)

    for r, c in enumerate(comments, start=2):
        ws.cell(row=r, column=1, value=c['id'])
        ws.cell(row=r, column=2, value=c['author'])
        ws.cell(row=r, column=3, value=c['date'])
        ws.cell(row=r, column=4, value=c['page_title'])
        ws.cell(row=r, column=5, value=c['page_url'])
        ws.cell(row=r, column=6, value=c['text'])

        if c['screenshot_url']:
            try:
                img_bytes = fetch(c['screenshot_url']).content
                png, w, h = normalize_png(img_bytes)
                path = os.path.join('pastel_screenshots', f"{c['id']}.png")
                with open(path, 'wb') as f:
                    f.write(png)
                xlimg = XLImage(io.BytesIO(png))
                fit_image(ws, xlimg, row=r, col=7, max_w_px=480, max_h_px=320, min_row_h_px=max(120, min(360, h)))
            except Exception as e:
                ws.cell(row=r, column=7, value=f'(screenshot fetch failed: {e})')

    wb.save(args.out)
    print(f'Done: {args.out} with {len(comments)} comments. Images in ./pastel_screenshots/')

if __name__ == '__main__':
    main()
