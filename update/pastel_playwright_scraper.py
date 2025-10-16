import argparse
import os
import io
import time
from playwright.sync_api import sync_playwright
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

def normalize_png(data):
    im = Image.open(io.BytesIO(data)).convert('RGBA')
    out = io.BytesIO()
    im.save(out, format='PNG')
    return out.getvalue(), im.width, im.height

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def fit_image(ws, xlimg, row, col, max_w_px=480, max_h_px=320, min_row_h_px=110):
    w, h = xlimg.width, xlimg.height
    scale = min(max_w_px / w, max_h_px / h, 1.0)
    xlimg.width = int(w * scale)
    xlimg.height = int(h * scale)
    ws.add_image(xlimg, f'{get_column_letter(col)}{row}')
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, min_row_h_px * 0.75)

def extract_comments_with_playwright(url):
    comments = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        print(f'Navigating to {url}...')
        page.goto(url)

        # Wait for comments to load
        print('Waiting for comments to load...')
        page.wait_for_timeout(3000)

        # Find all comment buttons in the sidebar
        comment_buttons = page.locator('button[class*="comment"]').all()
        total_comments = len(comment_buttons)
        print(f'Found {total_comments} comments')

        # Alternative: try finding by the pattern we see in the snapshot
        if total_comments == 0:
            # Look for buttons that contain author name and date pattern
            comment_buttons = page.locator('button:has-text("Andrew Ausel")').all()
            total_comments = len(comment_buttons)
            print(f'Found {total_comments} comments using alternative selector')

        for idx in range(total_comments):
            try:
                # Re-query comment buttons each time to avoid stale references
                comment_buttons = page.locator('button:has-text("Andrew Ausel")').all()
                if idx >= len(comment_buttons):
                    break

                button = comment_buttons[idx]

                # Get comment preview text
                button_text = button.inner_text()
                print(f'Processing comment {idx + 1}/{total_comments}: {button_text[:50]}...')

                # Click the comment to expand/view it
                button.click()
                page.wait_for_timeout(1000)  # Wait for comment to expand

                # Extract comment details from the expanded view
                # Look for the comment content area
                comment_content = ''
                author = 'Unknown'
                date = ''
                page_url = url

                # Try to extract from the button itself
                lines = button_text.split('\n')
                if len(lines) >= 2:
                    # First line usually has author and date
                    first_line = lines[0]
                    if 'Oct' in first_line or 'Sep' in first_line or 'Nov' in first_line:
                        parts = first_line.rsplit(' ', 2)  # Split from right to get date
                        if len(parts) >= 3:
                            author = parts[0]
                            date = ' '.join(parts[1:])
                    else:
                        author = first_line

                    # Rest is the comment
                    comment_content = '\n'.join(lines[1:])

                # Try to get page context from nearby elements
                page_title = ''
                try:
                    # Look for page path in the parent container
                    parent = button.locator('xpath=ancestor::*[contains(@class, "comment-group")]').first
                    if parent:
                        path_elem = parent.locator('text=/\\/articles\\//').first
                        if path_elem:
                            page_title = path_elem.inner_text()
                except:
                    pass

                # Try to capture screenshot if there's an image
                screenshot_path = None
                try:
                    # Look for image in the comment area
                    img = page.locator('img[src*="pastel"]').first
                    if img:
                        screenshot_path = f'pastel_screenshots/comment_{idx + 1}.png'
                        img.screenshot(path=screenshot_path)
                except:
                    pass

                comments.append({
                    'id': f'c{idx + 1}',
                    'author': author.strip(),
                    'date': date.strip(),
                    'page_title': page_title,
                    'page_url': page_url,
                    'text': comment_content.strip(),
                    'screenshot_path': screenshot_path
                })

            except Exception as e:
                print(f'Error processing comment {idx + 1}: {e}')
                continue

        browser.close()

    return comments

def main():
    ap = argparse.ArgumentParser(description='Export Pastel comments using Playwright automation.')
    ap.add_argument('url', help='Public Pastel share URL')
    ap.add_argument('--out', default='pastel_comments_playwright.xlsx', help='Output Excel filename')
    args = ap.parse_args()

    os.makedirs('pastel_screenshots', exist_ok=True)

    comments = extract_comments_with_playwright(args.url)

    if not comments:
        print('No comments found.')
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pastel Comments'

    headers = ['ID', 'Author', 'Date', 'Page Title', 'Page URL', 'Comment', 'Screenshot']
    ws.append(headers)

    # column widths
    set_col_width(ws, 1, 10)
    set_col_width(ws, 2, 20)
    set_col_width(ws, 3, 18)
    set_col_width(ws, 4, 42)
    set_col_width(ws, 5, 45)
    set_col_width(ws, 6, 90)
    set_col_width(ws, 7, 62)

    for r, c in enumerate(comments, start=2):
        ws.cell(row=r, column=1, value=c['id'])
        ws.cell(row=r, column=2, value=c['author'])
        ws.cell(row=r, column=3, value=c['date'])
        ws.cell(row=r, column=4, value=c['page_title'])
        ws.cell(row=r, column=5, value=c['page_url'])
        ws.cell(row=r, column=6, value=c['text'])

        if c['screenshot_path'] and os.path.exists(c['screenshot_path']):
            try:
                with open(c['screenshot_path'], 'rb') as f:
                    img_bytes = f.read()
                png, w, h = normalize_png(img_bytes)
                xlimg = XLImage(io.BytesIO(png))
                fit_image(ws, xlimg, row=r, col=7, max_w_px=480, max_h_px=320, min_row_h_px=max(120, min(360, h)))
            except Exception as e:
                ws.cell(row=r, column=7, value=f'(screenshot error: {e})')

    wb.save(args.out)
    print(f'Done: {args.out} with {len(comments)} comments. Images in ./pastel_screenshots/')

if __name__ == '__main__':
    main()
